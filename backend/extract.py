"""Extractor v3 — cell-anchored layouts with a 3-level hierarchy for the form:

    section  (true full-width / side-by-side title rows, e.g. "SM150 DRIVE STATUS")
      └ rowKey   (stand / device, e.g. "FM # 1")  -> collapsible in the UI
          └ subGroup (column-group header, e.g. "D M WATER", "OP PANEL",
                      "JACKING UNIT (DE SIDE)")
              └ colKey (leaf checkpoint, e.g. "CONDUCTIVITY")

Each input also gets a UI type (choice/number/text) + options.

Run:  python backend/extract.py   (regenerates schemas/ from the template)
"""
import openpyxl, json, re, os
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC = os.path.join(ROOT, "template", "MillShiftCheckList.xlsx")
OUT = os.path.join(ROOT, "schemas")
os.makedirs(OUT, exist_ok=True)

wb = openpyxl.load_workbook(SRC)
wbv = openpyxl.load_workbook(SRC, data_only=True)
LOCKED = {"na", "n/a", "-", "_", "--", "nil"}

idx = wbv["INDEX"]; desc = {}
for r in range(3, 16):
    sn = idx.cell(r, 2).value; d = idx.cell(r, 3).value
    if sn:
        m = re.search(r"(\d+)", str(sn))
        if m: desc[int(m.group(1))] = (str(d).strip() if d else "")

def txt(v): return "" if v is None else re.sub(r"\s+", " ", str(v)).strip()

def pretty(v):
    """De-letter-space decorative titles: 'S M  1 5 0  D R I V E' -> 'SM 150 DRIVE'.
    Words are separated by 2+ spaces; single-char runs within a word are joined."""
    if v is None: return ""
    words = re.split(r"\s{2,}", str(v).strip())
    out = []
    for w in words:
        toks = w.split()
        if toks and all(len(t) <= 1 for t in toks):
            out.append("".join(toks))
        elif toks:
            out.append(" ".join(toks))
    return " ".join(out)
def is_num_label(s): return bool(re.fullmatch(r"\d+", s.strip()))
def is_bordered(c):
    b = c.border
    return sum(1 for s in (b.left, b.right, b.top, b.bottom) if s and s.style) >= 2

def merge_maps(ws):
    cover = {}; anchor = {}; span = {}
    for mr in ws.merged_cells.ranges:
        a = (mr.min_row, mr.min_col)
        anchor[a] = (mr.max_row - mr.min_row + 1, mr.max_col - mr.min_col + 1)
        for r in range(mr.min_row, mr.max_row + 1):
            for cc in range(mr.min_col, mr.max_col + 1):
                span[(r, cc)] = a
                if (r, cc) != a: cover[(r, cc)] = a
    return cover, anchor, span

def infer_type(header_texts, section, rowkey):
    blob = " ".join(header_texts + [section, rowkey]).lower()
    h0 = (header_texts[0] if header_texts else "").lower()
    if "red" in blob and "green" in blob:   return "choice", ["Green", "Red"]
    if "green" in blob and "yellow" in blob: return "choice", ["Green", "Yellow"]
    if "ok/nok" in blob or "nok" in blob or "visual" in blob:
        return "choice", ["OK", "NOK", "NA"]
    units = ["°c", "deg c", "bar", "kω", "kΩ", "l/min", "µs", "us/cm", "%", "amp", "rpm", "humid", "/100"]
    measure_words = ["temp", "press", "flow", "level", "current", "conductivity", "ir value", "winding", "bearing"]
    if any(u in blob for u in units) or any(w in blob for w in measure_words):
        if any(w in h0 for w in ["status", "condition", "condn"]):
            return "choice", ["OK", "NOK", "NA"]
        return "number", None
    status_words = ["mounting", "leakage", "cable", "conduit", "jb", "tb", "signal", "door",
                    "cooling", "ventilation", "component", "cleanness", "house keeping", "condn",
                    "condition", "status", "insulation", "power supply", "foundation", "coupling",
                    "encoder", "blower", "filter", "sound", "leak", "fan", "light", "accumul", "heater"]
    if any(w in blob for w in status_words):
        return "choice", ["OK", "NOK", "NA"]
    return "text", None

all_sheets = []
for sid in range(1, 14):
    ws = wb[f"Sheet{sid}"]; cover, anchor, span = merge_maps(ws)
    maxr, maxc = ws.max_row, ws.max_column
    sheet_title = pretty(wbv[f"Sheet{sid}"]["A1"].value) or f"Sheet {sid}"

    # ----- detect title rows (sections). A title row's non-empty top-left cells
    # are ALL labels spanning >=2 cols, at least one spans >=3, and none are inputs.
    title_rows = {}
    for r in range(1, maxr + 1):
        spans = []; ok = True; has_wide = False
        for c in range(1, maxc + 1):
            if (r, c) in cover: continue
            v = txt(ws.cell(r, c).value)
            if v == "": continue
            rs, cs = anchor.get((r, c), (1, 1))
            if v.lower() in LOCKED or cs < 2: ok = False; break
            if cs >= 3: has_wide = True
            spans.append((c, c + cs - 1, pretty(ws.cell(r, c).value)))
        if ok and has_wide and spans:
            title_rows[r] = spans

    def cell_text(r, c):
        a = span.get((r, c), (r, c)); return txt(ws.cell(*a).value)

    def cell_pretty(r, c):
        a = span.get((r, c), (r, c)); return pretty(ws.cell(*a).value)

    def section_above(r, c):
        for tr in sorted([t for t in title_rows if t < r], reverse=True):
            for (c1, c2, t) in title_rows[tr]:
                if c1 <= c <= c2:
                    return t
        return sheet_title

    def headers_above(r, c):
        rr = r - 1
        while rr >= 1 and rr not in title_rows and cell_text(rr, c) == "":
            rr -= 1
        chain = []
        while rr >= 1 and rr not in title_rows:
            v = cell_pretty(rr, c)
            if v == "": break
            chain.append(v); rr -= 1
        return list(dict.fromkeys(chain))

    def equipment_left(r, c):
        for cc in range(c - 1, 0, -1):
            a = span.get((r, cc), (r, cc)); rs, cs = anchor.get(a, (1, 1))
            v = txt(ws.cell(*a).value)
            if v and rs > 1: return v
        return ""

    def rowkey_left(r, c):
        for cc in range(1, c):
            v = cell_text(r, cc)
            if v and not is_num_label(v): return v
        return ""

    cells = []; fields = []; ordn = 0
    for r in range(1, maxr + 1):
        for c in range(1, maxc + 1):
            if (r, c) in cover: continue
            cell = ws.cell(r, c); rs, cs = anchor.get((r, c), (1, 1))
            v = txt(cell.value); coord = f"{get_column_letter(c)}{r}"
            base = dict(r=r, c=c, coord=coord, rs=rs, cs=cs)
            if v:
                base.update(t=v, k=("locked" if v.lower() in LOCKED else "label"),
                            b=bool(cell.font and cell.font.bold))
                cells.append(base)
            elif is_bordered(cell) and rs == 1 and cs == 1 and r not in title_rows:
                chain = headers_above(r, c)
                colKey = chain[0] if chain else ""
                subGroup = chain[1] if len(chain) > 1 else ""
                section = section_above(r, c)
                equip = equipment_left(r, c)
                rk = rowkey_left(r, c)
                ft, opts = infer_type(chain, section, rk)
                fid = f"S{sid}_{coord}"
                fld = dict(id=fid, cell=coord, ord=ordn, section=section,
                           equipment=equip, subGroup=subGroup,
                           rowKey=rk, colKey=(colKey or subGroup), ft=ft)
                if opts: fld["opts"] = opts
                fields.append(fld); ordn += 1
                base.update(k="input", id=fid, ft=ft)
                if opts: base["opts"] = opts
                cells.append(base)
            else:
                base.update(k="empty"); cells.append(base)

    sheet = dict(sheetId=sid, name=f"Sheet{sid}", title=sheet_title,
                 description=desc.get(sid, ""), maxRow=maxr, maxCol=maxc,
                 merges=[[m.min_row, m.min_col, m.max_row, m.max_col] for m in ws.merged_cells.ranges],
                 colWidths={k: round(v.width, 1) for k, v in ws.column_dimensions.items() if v.width},
                 fieldCount=len(fields), cells=cells, fields=fields)
    json.dump(sheet, open(os.path.join(OUT, f"sheet{sid}.layout.json"), "w", encoding="utf-8"),
              ensure_ascii=False, indent=1)
    nsec = len({f["section"] for f in fields})
    all_sheets.append(dict(sheetId=sid, description=desc.get(sid, ""), fieldCount=len(fields)))
    print(f"Sheet{sid:>2}: {len(fields):>4} fields | {nsec:>2} sections | {desc.get(sid,'')[:38]}")

json.dump(dict(sheets=all_sheets), open(os.path.join(OUT, "index.json"), "w", encoding="utf-8"),
          ensure_ascii=False, indent=1)
print("TOTAL:", sum(s["fieldCount"] for s in all_sheets))
