"""Stamp a submission's values into a copy of the original Excel template.

Usage:  python stamp.py <template.xlsx> <output.xlsx>
        (payload JSON on stdin: {sheetId, date, shift, meta, values})

openpyxl edits cell values in place without touching styles, so the output is
byte-for-byte the original formatting with the readings filled in.
"""
import sys, json, re
import openpyxl
from openpyxl.utils import range_boundaries, get_column_letter

def main():
    template, out = sys.argv[1], sys.argv[2]
    payload = json.load(sys.stdin)
    sid = int(payload["sheetId"])
    values = payload.get("values", {})
    meta = payload.get("meta", {})

    wb = openpyxl.load_workbook(template)
    ws = wb[f"Sheet{sid}"]
    # Hide non-target sheets to keep the downloaded file single-sheet focused
    # without breaking any underlying merged cells, print areas, or row dimensions
    for sn in wb.sheetnames:
        if sn != f"Sheet{sid}":
            wb[sn].sheet_state = 'hidden'

    # Ensure the target sheet is active
    wb.active = wb[f"Sheet{sid}"]
    # map of any merged cell -> its top-left anchor coordinate
    anchor = {}
    for mr in ws.merged_cells.ranges:
        c1, r1, c2, r2 = mr.min_col, mr.min_row, mr.max_col, mr.max_row
        a = f"{get_column_letter(c1)}{r1}"
        for r in range(r1, r2 + 1):
            for c in range(c1, c2 + 1):
                anchor[f"{get_column_letter(c)}{r}"] = (a, c2)

    # ---- stamp the readings (keys are exact cell addresses) ----
    for cell, val in values.items():
        if val is None or val == "":
            continue
        try:
            ws[cell] = val
        except Exception:
            pass  # skip anything that lands on a protected/merged child

    # ---- highlight anomalies ----
    anomalies = meta.get("anomalies", [])
    if anomalies:
        from copy import copy
        for a in anomalies:
            acell = a.get("cell")
            if acell:
                try:
                    c = ws[acell]
                    if c.font:
                        new_font = copy(c.font)
                        new_font.color = "FF0000"
                        new_font.bold = True
                        c.font = new_font
                    else:
                        from openpyxl.styles import Font
                        c.font = Font(color="FF0000", bold=True)
                except Exception:
                    pass

    # ---- best-effort header metadata ----
    def stamp_after(keyword, value):
        if not value:
            return
        kw = keyword.lower()
        for row in ws.iter_rows():
            for c in row:
                t = c.value
                norm = re.sub(r"\s+", " ", str(t)).strip().lower() if t else ""
                # only a genuine field label: "Date :", "SHIFT:", "CHECKED BY:" —
                # starts with the keyword and carries a colon. Avoids "Issue Date".
                if norm.startswith(kw) and ":" in norm:
                    # find right edge of this label (if merged) then write to next cell
                    coord = c.coordinate
                    end_col = c.column
                    if coord in anchor:
                        end_col = anchor[coord][1]
                    target = f"{get_column_letter(end_col + 1)}{c.row}"
                    tgt = anchor.get(target, (target, None))[0]
                    try:
                        ws[tgt] = value
                    except Exception:
                        pass
                    return

    who = meta.get("checkedBy", "")
    if meta.get("empId"):
        who = f"{who} ({meta['empId']})".strip()
    stamp_after("date", payload.get("date", ""))
    stamp_after("shift", payload.get("shift", ""))
    stamp_after("checked by", who)

    wb.save(out)
    print(f"stamped Sheet{sid}: {len(values)} values -> {out}", file=sys.stderr)

if __name__ == "__main__":
    main()
